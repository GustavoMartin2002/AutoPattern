import io
from datetime import datetime, timedelta, timezone
from typing import Any

import pandas as pd

from core.domain.exceptions import ReportGenerationError
from core.domain.interfaces.report_generator import IReportGenerator


class ReportAdapter(IReportGenerator):
  """
  Adaptador para geração de relatórios em Excel e PDF.
  Funcionalidades:
  - Geração de planilhas Excel com formatação profissional
  - Geração de PDFs com resumo executivo inteligente
  - Limpeza automática de dados para gráficos
  - Aplicação de timezone do Brasil (UTC-3)
  """

  def _prepare_chart_data(
    self, labels: Any, values: Any
  ) -> tuple[list[str], list[float]]:
    """Prepara e limpa dados para uso em gráficos."""
    clean_labels = []
    clean_values = []

    for label, value in zip(labels, values, strict=False):
      try:
        val_float = float(value) if pd.notnull(value) else 0.0
      except (ValueError, TypeError):
        val_float = 0.0

      if val_float > 0:
        clean_labels.append(str(label))
        clean_values.append(val_float)

    return clean_labels, clean_values

  def generate_excel(self, data: dict[str, Any]) -> bytes:
    """Gera arquivo Excel (.xlsx) com dados hierárquicos."""
    try:
      from openpyxl import Workbook
      from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
      from openpyxl.utils.dataframe import dataframe_to_rows

      items = data.get("items", [])
      if not items:
        # Caso especial: XML vazio ou sem dados
        wb = Workbook()
        ws = wb.active
        ws.title = "AutoPattern"
        ws["A1"] = "Nenhum dado encontrado"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

      # Converte dados para DataFrame do pandas
      df = pd.DataFrame(items)
      df = df.fillna("")  # Substitui valores nulos por string vazia

      wb = Workbook()
      ws = wb.active
      ws.title = "AutoPattern"

      # Define estilos de formatação
      header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
      header_fill = PatternFill(
        start_color="2563eb", end_color="2563eb", fill_type="solid"
      )
      center_aligned = Alignment(
        horizontal="center", vertical="center", wrap_text=False
      )
      thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
      )

      if not df.empty:
        # Escreve dados no Excel (cabeçalho + linhas)
        for r_idx, row in enumerate(
          dataframe_to_rows(df, index=False, header=True), 1
        ):
          for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_aligned
            cell.border = thin_border

            # Aplica estilo especial para cabeçalho (primeira linha)
            if r_idx == 1:
              cell.font = header_font
              cell.fill = header_fill

        # Ajusta largura das colunas automaticamente
        for column_cells in ws.columns:
          length = max(len(str(cell.value)) for cell in column_cells)
          width = min(length + 4, 50)  # Máximo de 50 caracteres
          ws.column_dimensions[column_cells[0].column_letter].width = width

      output = io.BytesIO()
      wb.save(output)
      return output.getvalue()
    except Exception as e:
      raise ReportGenerationError(
        f"Erro ao gerar relatório Excel: {e!s}"
      ) from e

  def generate_pdf(self, data: dict[str, Any]) -> bytes:
    """Gera PDF com Resumo Executivo Inteligente."""
    try:
      from reportlab.lib import colors
      from reportlab.lib.enums import TA_CENTER
      from reportlab.lib.pagesizes import A4
      from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
      from reportlab.lib.units import cm
      from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
      )

      buffer = io.BytesIO()
      doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm
      )

      elementos = []
      styles = getSampleStyleSheet()

      # ESTILOS PERSONALIZADOS
      titulo_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#ffffff"),
        spaceAfter=30,
        alignment=TA_CENTER,
      )

      subtitulo_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=16,
        textColor=colors.HexColor("#2563eb"),
        spaceAfter=12,
        spaceBefore=20,
      )

      tag_name_style = ParagraphStyle(
        "TagName",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#7f8c8d"),
        fontName="Helvetica-Bold",
      )

      tag_value_style = ParagraphStyle(
        "TagValue",
        parent=styles["Normal"],
        fontSize=12,
        textColor=colors.HexColor("#2c3e50"),
      )

      # CABEÇALHO DESTACADO
      header_data = [
        [
          Paragraph("Relatório de Dados XML", titulo_style),
        ]
      ]

      header_table = Table(header_data, colWidths=[16 * cm])
      header_table.setStyle(
        TableStyle(
          [
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 20),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
          ]
        )
      )
      elementos.append(header_table)
      elementos.append(Spacer(1, 0.5 * cm))

      # DATA DE GERAÇÃO (HORÁRIO DO BRASIL)
      brazil_tz = timezone(timedelta(hours=-3))
      now_brazil = datetime.now(brazil_tz)
      data_geracao = Paragraph(
        f"<i>Gerado em: {now_brazil.strftime('%d/%m/%Y às %H:%M')}</i>",
        styles["Normal"],
      )
      elementos.append(data_geracao)
      elementos.append(Spacer(1, 1 * cm))

      # ESTATÍSTICAS EM CARDS
      stats = data.get(
        "statistics", {"total": 0, "preenchidas": 0, "vazias": 0}
      )
      stats_data = [
        ["Total de Tags", "Tags Preenchidas", "Tags Vazias"],
        [
          str(stats.get("total", 0)),
          str(stats.get("preenchidas", 0)),
          str(stats.get("vazias", 0)),
        ],
      ]

      stats_table = Table(stats_data, colWidths=[5.3 * cm, 5.3 * cm, 5.3 * cm])
      stats_table.setStyle(
        TableStyle(
          [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("TOPPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("FONTSIZE", (0, 1), (-1, -1), 18),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 1), (-1, -1), 15),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 15),
            ("GRID", (0, 0), (-1, -1), 1, colors.white),
          ]
        )
      )
      elementos.append(stats_table)
      elementos.append(Spacer(1, 1 * cm))

      # TAGS PRIORITÁRIAS (MÁXIMO 10)
      elementos.append(Paragraph("Top 5 Tags Mais Frequentes", subtitulo_style))
      elementos.append(Spacer(1, 0.4 * cm))

      tags_criticas = data.get("tags_criticas", [])
      if tags_criticas:
        elementos.append(Paragraph("Informações Prioritárias", subtitulo_style))
        elementos.append(Spacer(1, 0.3 * cm))

        # Limita a 10 tags prioritárias para controlar tamanho do PDF
        for tag in tags_criticas[:10]:
          card_data = [
            [Paragraph(tag["nome"], tag_name_style)],
            [Paragraph(str(tag["valor"]), tag_value_style)],
          ]

          card_table = Table(card_data, colWidths=[16 * cm])
          card_table.setStyle(
            TableStyle(
              [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8f9fa")),
                ("LEFTPADDING", (0, 0), (-1, -1), 15),
                ("RIGHTPADDING", (0, 0), (-1, -1), 15),
                ("TOPPADDING", (0, 0), (0, 0), 10),
                ("BOTTOMPADDING", (0, 1), (0, 1), 10),
                ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#2563eb")),
              ]
            )
          )

          elementos.append(card_table)
          elementos.append(Spacer(1, 0.3 * cm))

      # AGRUPAMENTOS (MÁXIMO 5 GRUPOS)
      grupos = data.get("grupos", {})
      if grupos:
        # Limita a 5 grupos para controlar tamanho do PDF
        for grupo_nome, tags in list(grupos.items())[:5]:
          elementos.append(Spacer(1, 0.5 * cm))
          elementos.append(Paragraph(f"{grupo_nome}", subtitulo_style))
          elementos.append(Spacer(1, 0.2 * cm))

          # Limita a 10 itens por grupo
          grupo_data = [
            [
              Paragraph(tag["nome"], tag_name_style),
              Paragraph(str(tag["valor"])[:50] + "...", tag_value_style)
              if len(str(tag["valor"])) > 50
              else Paragraph(str(tag["valor"]), tag_value_style),
            ]
            for tag in tags[:10]
          ]

          # Se há mais de 10 itens, mostra mensagem
          if len(tags) > 10:
            grupo_data.append(
              [
                Paragraph(
                  f"<i>... e mais {len(tags) - 10} itens</i>", styles["Normal"]
                ),
                Paragraph(
                  "<i>(consulte o Excel para detalhes completos)</i>",
                  styles["Normal"],
                ),
              ]
            )

          grupo_table = Table(grupo_data, colWidths=[7 * cm, 9 * cm])
          grupo_table.setStyle(
            TableStyle(
              [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
              ]
            )
          )
          elementos.append(grupo_table)

      # ALERTAS
      alerts = data.get("alerts", [])
      if alerts:
        elementos.append(Spacer(1, 1 * cm))
        elementos.append(Paragraph("Alertas", subtitulo_style))
        for alert in alerts:
          elementos.append(Paragraph(alert, styles["Normal"]))
          elementos.append(Spacer(1, 0.2 * cm))

      # Gera o PDF
      doc.build(elementos)
      return buffer.getvalue()
    except Exception as e:
      raise ReportGenerationError(f"Erro ao gerar relatório PDF: {e!s}") from e
